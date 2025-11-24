import pandas as pd
import numpy as np
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

def GradeThreshold(summary_df, CV_High, CV_Low, excel_save_path):
    required_cols = ["工位", "更新时间", "结果"]
    if not all(col in summary_df.columns for col in required_cols):
        raise ValueError(f"summary_df 必须包含以下列:{required_cols}")
    if summary_df.empty:
        warnings.warn("summary_df 为空，无数据可处理")
        return pd.DataFrame(), {}

    summary_df = summary_df.copy()
    all_times = sorted(summary_df["更新时间"].unique())
    print(f"===== 分级阈值计算(共 {len(all_times)} 个时间段)=====\n")

    time_thresholds = {}
    historical_data = pd.DataFrame()
    summary_df["等级"] = None

    for current_time in all_times:
        time_data = summary_df[summary_df["更新时间"] == current_time].copy()
        time_values = time_data["结果"].values
        n_stations = len(time_data)
        if n_stations == 0:
            warnings.warn(f"时间段 {current_time} 无工位数据，跳过")
            continue

        mu_time = np.mean(time_values)
        sigma_time = np.std(time_values)
        cv_time = sigma_time / abs(mu_time) if mu_time != 0 else 0

        if not historical_data.empty:
            historical_values = historical_data["结果"].values
            mu_historical = np.mean(historical_values)
            sigma_historical = np.std(historical_values)
            has_history = True
        else:
            mu_historical = mu_time
            sigma_historical = sigma_time
            has_history = False
            warnings.warn(f"时间段 {current_time} 是第一个时间段，无历史数据，使用本时间段数据作为参考")

        if cv_time < CV_Low:
            threshold_type = "历史统计量(CV < CV_Low)"
            T_high = mu_historical + 0.5 * sigma_historical
            T_low = mu_historical - 0.5 * sigma_historical
        elif cv_time > CV_High:
            threshold_type = "本时间段统计量(CV > CV_High)"
            T_high = mu_time + 0.8 * sigma_time
            T_low = mu_time - 0.8 * sigma_time
        else:
            threshold_type = "本时间段分位数（常规）"
            T_high = np.quantile(time_values, 0.75)
            T_low = np.quantile(time_values, 0.25)

        def get_grade(value):
            if value <= T_low:
                return "中"
            elif value >= T_high:
                return "优"
            else:
                return "良"

        summary_df.loc[summary_df["更新时间"] == current_time, "等级"] = summary_df.loc[summary_df["更新时间"] == current_time, "结果"].apply(get_grade)
        summary_df.loc[summary_df["更新时间"] == current_time, "T_high"] = T_high
        summary_df.loc[summary_df["更新时间"] == current_time, "T_low"] = T_low


        time_thresholds[current_time] = {
            "时间段": current_time,
            "工位数": n_stations,
            "本时间段均值(μ_t)": round(mu_time, 4),
            "本时间段标准差(σ_t)": round(sigma_time, 4),
            "本时间段CV": round(cv_time, 4),
            "是否有历史数据": has_history,
            "历史数据量": len(historical_data) if has_history else 0,
            "历史均值(μ_历史)": round(mu_historical, 4) if has_history else "无",
            "历史标准差(σ_历史)": round(sigma_historical, 4) if has_history else "无",
            "阈值计算方式": threshold_type,
            "T_high": round(T_high, 8),
            "T_low": round(T_low, 8)
        }

        historical_data = pd.concat([historical_data, time_data], ignore_index=True)

        print(f"=== 时间段 {current_time} ===")
        print(f"工位数: {n_stations} | 本时间段CV: {cv_time:.4f} | 阈值方式: {threshold_type}")
        print(f"历史数据量: {len(historical_data) - n_stations} (截至上一时间段)")
        print(f"T_high: {T_high:.8f} | T_low: {T_low:.8f}")
        print(f"等级分布: {summary_df[summary_df['更新时间'] == current_time]['等级'].value_counts().to_dict()}\n")

    # 添加结论列
    def generate_conclusion(row):
        if row['等级'] == '优':
            return ""
        weighted_cols = ['检验成本_加权值', '不合格率_加权值', '返工成本_加权值', '报废成本_加权值']
        values = {col: row[col] for col in weighted_cols if pd.notnull(row[col])}
        sorted_items = sorted(values.items(), key=lambda x: x[1], reverse=True)
        if row['等级'] == '良' and sorted_items:
            col, val = sorted_items[0]
            return f"{col}={val}"
        elif row['等级'] == '中' and len(sorted_items) >= 2:
            (col1, val1), (col2, val2) = sorted_items[:2]
            return f"{col1}={val1}，{col2}={val2}"
        elif row['等级'] == '中' and sorted_items:
            col, val = sorted_items[0]
            return f"{col}={val}"
        return ""

    summary_df['结论'] = summary_df.apply(generate_conclusion, axis=1)

    result_df = summary_df.copy()

    print("===== 所有时间段处理完成 =====")
    print(f"有效时间段数:{len(time_thresholds)}")
    print(f"总打分记录数:{len(result_df)}")
    if not result_df.empty:
        print(f"全局等级分布:{result_df['等级'].value_counts().to_dict()}")

    try:
        result_df.to_excel(excel_save_path, index=False, engine="openpyxl")

        wb = load_workbook(excel_save_path)
        ws = wb.active
        ws.title = "工位等级结果"

        green_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        red_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")

        grade_col_idx = list(result_df.columns).index("等级") + 1
        col_letter = chr(64 + grade_col_idx) if grade_col_idx <= 26 else 'Z'
        data_range = f"{col_letter}2:{col_letter}{len(result_df) + 1}"

        rule_good = CellIsRule(operator="equal", formula=['"优"'], fill=green_fill)
        rule_bad = CellIsRule(operator="equal", formula=['"中"'], fill=red_fill)

        ws.conditional_formatting.add(data_range, rule_good)
        ws.conditional_formatting.add(data_range, rule_bad)

        wb.save(excel_save_path)
        print(f"\n✅ Excel已生成! 路径:{excel_save_path}")
        print(f"格式说明：优=浅绿底，中=浅红底，良=无底色")

    except Exception as e:
        raise RuntimeError(f"Excel生成失败:{str(e)}")

    return result_df, time_thresholds